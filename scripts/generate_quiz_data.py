#!/usr/bin/env python3
"""Narzędzie admina: neuro_questions.xlsx -> public/data/index.json + sections/<id>.json.
Logika czytania oparta na zewnętrznym (legacy) xlsx_to_csv.py z katalogu nadrzędnego
(poza VCS tego repo) — wyjście tutaj to JSON zamiast CSV."""
import datetime
import json
import re
import shutil
import sys
import unicodedata
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = REPO_ROOT / "public" / "data"

PL_MAP = str.maketrans({
    "ą": "a", "ć": "c", "ę": "e", "ł": "l", "ń": "n", "ó": "o", "ś": "s", "ź": "z", "ż": "z",
})

LETTER_TO_INDEX = {"A": 0, "B": 1, "C": 2, "D": 3}


def slugify(text: str) -> str:
    lowered = text.lower().translate(PL_MAP)
    decomposed = unicodedata.normalize("NFD", lowered)
    ascii_only = "".join(c for c in decomposed if unicodedata.category(c) != "Mn")
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", ascii_only)).strip("-")


def cell_text(cell) -> str:
    """Odwraca autokonwersje Excela (liczby, daty). Logika wg zewnętrznego (legacy) xlsx_to_csv.py."""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        fmt = (cell.number_format or "").lower()
        if "y" not in fmt:  # 'm-d' => wpisano "1-2", Excel zrobił datę
            return f"{v.month}-{v.day}"
        return v.date().isoformat()
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def parse_tags(raw: str) -> list[str]:
    """Forma hashowa lub po przecinku, lowercase, dedup. Logika wg zewnętrznego (legacy) xlsx_to_csv.py."""
    if not raw:
        return []
    hashed = re.findall(r"#([^\s#,]+)", raw)
    parts = [t.strip(" .,;") for t in hashed] if hashed else [t.strip() for t in raw.split(",")]
    return list(dict.fromkeys(t.lower() for t in parts if t))


def cites_letter(text: str) -> bool:
    """Wykrywa jawne cytowanie litery w wyjaśnieniu (wzorzec wąski, bez fałszywych alarmów)."""
    if not text:
        return False
    return bool(
        re.search(r"\bOdpowiedź [A-D]\b", text)
        or re.search(r"\b[A-D] jest (poprawn|nieprawidłow|błędn)", text)
    )


def is_excluded(raw: str) -> bool:
    """Kolumna 'Wykluczone': pytanie wadliwe merytorycznie, pomijane przy budowie danych aplikacji."""
    return raw.strip().lower() in {"tak", "true", "1", "x", "yes"}


def build_section_meta(section: dict) -> dict:
    topics = []
    for t in section["topics"]:
        count = sum(1 for q in section["questions"] if q["topic"] == t["title"])
        topics.append({"slug": t["slug"], "title": t["title"], "questionCount": count})
    return {
        "id": section["id"],
        "slug": slugify(section["title"]),
        "title": section["title"],
        "questionCount": len(section["questions"]),
        "file": f"sections/{section['id']}.json",
        "topics": topics,
    }


def main() -> None:
    import openpyxl  # tylko generacja potrzebuje xlsx; testy czystych funkcji (Task 1) nie

    if len(sys.argv) < 2:
        sys.exit("Użycie: python3 scripts/generate_quiz_data.py <ścieżka.xlsx>")
    xlsx_path = Path(sys.argv[1])
    wb = openpyxl.load_workbook(xlsx_path)  # bez data_only: number_format potrzebny
    struktura = wb["Struktura"]
    sheets = [s for s in wb.sheetnames if s != "Struktura"]

    # Struktura: (numer, dział, temat) — czytamy aż do pustego wiersza (bez hardkodu liczby).
    meta = []
    r = 1
    while True:
        num = struktura.cell(r, 1).value
        if num is None or str(num).strip() == "":
            break
        meta.append((int(num), cell_text(struktura.cell(r, 2)), cell_text(struktura.cell(r, 3))))
        r += 1
    if len(meta) != len(sheets):
        sys.exit(f"Niezgodność: {len(meta)} wierszy Struktury vs {len(sheets)} arkuszy pytań")

    errors: list[str] = []
    letter_warnings: list[str] = []
    excluded_count = 0
    seen_ids: set[str] = set()
    sections_by_id: dict[int, dict] = {}

    for (section_no, section_name, topic), sheet in zip(meta, sheets):
        prefix = re.match(r"^(\d+)\.", sheet)
        if not prefix or int(prefix.group(1)) != section_no:
            sys.exit(f"Arkusz '{sheet}' nie pasuje do działu {section_no} ze Struktury")
        ws = wb[sheet]
        topic_slug = slugify(topic)
        section = sections_by_id.setdefault(
            section_no, {"id": section_no, "title": section_name, "topics": [], "questions": []}
        )
        section["topics"].append({"slug": topic_slug, "title": topic})

        for row in range(2, ws.max_row + 1):
            cells = [ws.cell(row, c) for c in range(1, 11)]
            vals = [cell_text(c) for c in cells]
            if not any(vals):
                continue  # pusty wiersz
            _, question, a, b, c_, d, correct, explanation, tags, excluded = vals
            qid = f"{section_no}-{topic_slug}-{row}"

            if question == "Pytanie" and a == "A":
                continue  # powtórzony nagłówek (jak w xlsx_to_csv.py)
            if question and not any([a, b, c_, d]):
                continue  # wiersz-separator / tytuł bloku (jak w xlsx_to_csv.py)
            if is_excluded(excluded):
                excluded_count += 1
                continue  # pytanie wadliwe merytorycznie — nie trafia do quizu
            if not all([question, a, b, c_, d]):
                errors.append(f"[{qid}] niekompletne pytanie")
                continue
            if correct.upper() not in LETTER_TO_INDEX:
                errors.append(f"[{qid}] nieprawidłowa Poprawna={correct!r}")
                continue
            if cites_letter(explanation):
                letter_warnings.append(qid)
            if qid in seen_ids:
                errors.append(f"[{qid}] zduplikowane id")
                continue
            seen_ids.add(qid)

            section["questions"].append({
                "id": qid,
                "section": section_name,
                "topic": topic,
                "question": question,
                "options": [a, b, c_, d],
                "correctIndex": LETTER_TO_INDEX[correct.upper()],
                "explanation": explanation,
                "tags": parse_tags(tags),
            })

    if errors:
        print(f"WALIDACJA NIEUDANA — {len(errors)} błędów:", file=sys.stderr)
        for e in errors[:50]:
            print("  " + e, file=sys.stderr)
        if len(errors) > 50:
            print(f"  ...i {len(errors) - 50} więcej", file=sys.stderr)
        sys.exit(1)

    sections = [sections_by_id[k] for k in sorted(sections_by_id)]
    manifest = {
        "generatedFrom": xlsx_path.name,
        "totalQuestions": sum(len(s["questions"]) for s in sections),
        "sections": [build_section_meta(s) for s in sections],
    }

    sections_dir = OUT_DIR / "sections"
    if sections_dir.exists():
        shutil.rmtree(sections_dir)  # usuń osierocone pliki działów z poprzednich importów
    sections_dir.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "index.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    for s in sections:
        data = {"id": s["id"], "title": s["title"], "questions": s["questions"]}
        (sections_dir / f"{s['id']}.json").write_text(
            json.dumps(data, ensure_ascii=False) + "\n", encoding="utf-8"
        )

    total_topics = sum(len(s["topics"]) for s in sections)
    print(f"OK: {manifest['totalQuestions']} pytań, {len(sections)} działów, {total_topics} tematów")
    if excluded_count:
        print(f"Pominięto {excluded_count} pytań oznaczonych w kolumnie 'Wykluczone'")
    if letter_warnings:
        print(
            f"OSTRZEŻENIE: {len(letter_warnings)} wyjaśnień cytuje litery "
            f"(przepisz — spec sekcja 6). Np. {', '.join(letter_warnings[:5])}"
        )


if __name__ == "__main__":
    main()
