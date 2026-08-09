# -*- coding: utf-8 -*-
"""Oznacza jako wykluczone te kopie zdublowanych pytań, które siedzą w niewłaściwym temacie.
Wiersz zostaje w pliku — wykluczenie to flaga w kolumnie J, więc decyzja jest odwracalna."""
import json
import openpyxl

XLSX = '/Users/adamdrzewiecki/IntelliJProjects/private/neuro_quiz/neuro-wiry-hub/data/neuro_questions.xlsx'
LOG = '/Users/adamdrzewiecki/IntelliJProjects/private/neuro_quiz/neuro-wiry-hub/docs/validation/applied-changes.json'
COL_EXCL = 10

# (arkusz do wykluczenia, wiersz, arkusz zachowany, wiersz, uzasadnienie wyboru)
DECISIONS = [
    ('1.Neurohistologia', 16, '1.Ośrodki i drogi układu nerwow', 4,
     'pytanie o drogi kojarzeniowe należy do tematu o ośrodkach i drogach, nie do histologii'),
    ('1.Neurohistologia', 23, '1.Rozwój układu nerwowego', 23,
     'wodogłowie wrodzone to zagadnienie rozwojowe, nie histologiczne'),
    ('9.Znaczenie kliniczne', 17, '1.Rozwój układu nerwowego', 23,
     'trzecia kopia tego samego pytania o wodogłowie wrodzone'),
    ('1.Rozwój układu nerwowego', 25, '2.Rozwój mózgowia', 14,
     'pęcherzyki oczne to uwypuklenia międzymózgowia — rozwój mózgowia jest tematem właściwym'),
    ('1.Rozwój układu nerwowego', 30, '2.Rozwój mózgowia', 23,
     'powstawanie wzgórza z pęcherzyka wtórnego to rozwój mózgowia'),
    ('4.Płaty mózgu', 22, '1.Ośrodki i drogi układu nerwow', 15,
     'pytanie o drogi kojarzeniowe długie dotyczy dróg, nie podziału na płaty'),
    ('5.Wzgórze', 21, '2.Unaczynienie', 18,
     'pytanie o unaczynienie wzgórza należy do tematu o unaczynieniu'),
    ('8.Zakręt obręczy', 18, '8.Przegroda', 8,
     'pole podspoidłowe wchodzi w skład okolicy przegrody'),
    ('9.Połączenia jąder podstawy', 15, '11.Torebka wewnętrzna', 9,
     'pytanie o odnogę przednią torebki wewnętrznej należy do tematu o torebce'),
    ('11.Torebka wewnętrzna', 23, '1.Ośrodki i drogi układu nerwow', 24,
     'pytanie o zakończenie drogi słuchowej dotyczy przebiegu drogi, nie torebki'),
    ('14.Drogi wstępujące', 28, '7.Drogi móżdżkowe', 24,
     'droga rdzeniowo-móżdżkowa przednia jest opisana w temacie dróg móżdżkowych'),
    ('17.Drogi pozapiramidowe', 11, '10.Drogi pozapiramidowe', 3,
     'dział 10 jest dedykowany układowi pozapiramidowemu; kopia w dziale 17 jest wtórna'),
]


def main():
    wb = openpyxl.load_workbook(XLSX)
    log = json.load(open(LOG))
    done = 0
    for sheet, row, keep_sheet, keep_row, why in DECISIONS:
        ws = wb[sheet]
        q = str(ws.cell(row, 2).value or '')
        keep_q = str(wb[keep_sheet].cell(keep_row, 2).value or '')
        if q.strip().lower() != keep_q.strip().lower():
            print(f'  POMINIĘTO {sheet} w.{row}: treść nie zgadza się z kopią zachowywaną')
            print(f'    {q[:70]!r}\n    {keep_q[:70]!r}')
            continue
        if str(ws.cell(row, COL_EXCL).value or '').strip().lower() == 'tak':
            print(f'  {sheet} w.{row} już wykluczone')
            continue
        ws.cell(row, COL_EXCL).value = 'TAK'
        done += 1
        log.append({'sheet': sheet, 'row': row, 'id': str(ws.cell(row, 1).value), 'field': 'Wykluczone',
                    'from': '', 'to': 'TAK',
                    'why': f'duplikat pytania z {keep_sheet} w.{keep_row}: {why}',
                    'signals': 'deduplikacja'})
        print(f'  wykluczono {sheet} w.{row}  (zostaje {keep_sheet} w.{keep_row})')
    wb.save(XLSX)
    json.dump(log, open(LOG, 'w'), ensure_ascii=False, indent=1)
    print(f'\nwykluczonych duplikatów: {done}/{len(DECISIONS)} | dziennik: {len(log)}')


if __name__ == '__main__':
    main()
