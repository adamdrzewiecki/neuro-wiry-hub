# -*- coding: utf-8 -*-
"""Eksportuje 106 arkuszy xlsx do JSON-ów per temat (jeden plik = jeden arkusz)."""
import openpyxl, json, os, unicodedata, re, datetime

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
# Katalog roboczy: NIE w repo (pliki pośrednie są duże i regenerowalne).
# Nadpisz zmienną NEURO_WORK_DIR, żeby wskazać inne miejsce.
SCR = os.environ.get('NEURO_WORK_DIR') or os.path.join(REPO, '.validation-work')
XLSX = os.path.join(REPO, 'data', 'neuro_questions.xlsx')
INDEX = os.path.join(REPO, 'public', 'data', 'index.json')
OUT = os.path.join(SCR, 'topics')


def slugify(s):
    s = unicodedata.normalize('NFKD', str(s).replace('ł', 'l').replace('Ł', 'L'))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'[^A-Za-z0-9]+', '-', s).strip('-').lower()


def main():
    os.makedirs(OUT, exist_ok=True)
    print(f'katalog roboczy: {SCR}')
    idx = json.load(open(INDEX))
    topics_by_section = {s['id']: list(s['topics']) for s in idx['sections']}
    sec_title = {s['id']: s['title'] for s in idx['sections']}

    wb = openpyxl.load_workbook(XLSX, read_only=True)
    counters, manifest, weird, total = {}, [], [], 0
    for sheet in wb.sheetnames:
        if sheet == 'Struktura':
            continue
        sec_id = int(sheet.split('.', 1)[0])
        i = counters.get(sec_id, 0)
        counters[sec_id] = i + 1
        topic = topics_by_section[sec_id][i]
        ws = wb[sheet]
        rows = []
        for rn, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r[0] is None and r[1] is None:
                continue
            vals = list(r[:8])
            for ci, v in enumerate(vals):
                if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
                    weird.append(f'{sheet} w.{rn} kol.{ci}: {v!r}')
                    vals[ci] = str(v)
            rows.append({'row': rn, 'id': vals[0], 'question': vals[1], 'A': vals[2], 'B': vals[3],
                         'C': vals[4], 'D': vals[5], 'correct': vals[6], 'explanation': vals[7]})
        fname = f'{sec_id:02d}-{slugify(topic["title"])}.json'
        json.dump({'sheet': sheet, 'sectionId': sec_id, 'sectionTitle': sec_title[sec_id],
                   'topicTitle': topic['title'], 'questionCount': len(rows), 'questions': rows},
                  open(os.path.join(OUT, fname), 'w'), ensure_ascii=False, indent=1)
        manifest.append({'file': f'topics/{fname}', 'sheet': sheet, 'sectionId': sec_id,
                         'sectionTitle': sec_title[sec_id], 'topicTitle': topic['title'], 'count': len(rows)})
        total += len(rows)

    json.dump(manifest, open(os.path.join(SCR, 'topics-manifest.json'), 'w'), ensure_ascii=False, indent=1)
    print(f'arkuszy: {len(manifest)} | pytań: {total} | index.json deklaruje: {idx["totalQuestions"]}')
    print(f'komórki daty (uszkodzone przez Excel): {weird}')


if __name__ == '__main__':
    main()
