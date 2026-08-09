# -*- coding: utf-8 -*-
"""Szuka znaków spoza alfabetu łacińskiego w treści pytań i opcji (np. wklejone znaki CJK)."""
import openpyxl
import re
import unicodedata

XLSX = '/Users/adamdrzewiecki/IntelliJProjects/private/neuro_quiz/neuro-wiry-hub/data/neuro_questions.xlsx'
CJK = re.compile(r'[　-鿿가-힯Ѐ-ӿ]')

wb = openpyxl.load_workbook(XLSX, read_only=True)
hits = []
for sh in wb.sheetnames:
    if sh == 'Struktura':
        continue
    for i, r in enumerate(wb[sh].iter_rows(min_row=2, values_only=True), start=2):
        if r[0] is None and r[1] is None:
            continue
        for j, name in enumerate(['ID', 'Pytanie', 'A', 'B', 'C', 'D', 'Poprawna', 'Wyjaśnienie']):
            v = str(r[j] or '')
            m = CJK.search(v)
            if m:
                ch = m.group(0)
                hits.append((sh, i, name, ch, unicodedata.name(ch, '?'), v[:70]))

print(f'znalezionych komórek ze znakami spoza łacinki: {len(hits)}')
for h in hits:
    print(f'  {h[0]} w.{h[1]} {h[2]}: {h[3]!r} ({h[4]}) w tekście: {h[5]!r}')

ws = wb['14.Drogi zstępujące']
print('\n--- 14.Drogi zstępujące, wiersz 13 ---')
for r in ws.iter_rows(min_row=13, max_row=13, values_only=True):
    print(' pytanie:', r[1])
    for i, k in enumerate('ABCD'):
        print(f'  {k}: {r[2 + i]!r}')
    print(' klucz:', r[6])
